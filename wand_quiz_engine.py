# wand_quiz_engine.py
# A wrapper that exposes your Pottermore spreadsheet logic
# without printing or using input().
#
# This lets your game call it using your own UI system.

from openpyxl import load_workbook

EXCEL_FILE = "Pottermore Wand Selection Spreadsheet.xlsx"

# -----------------------------
# DATA LOADING
# -----------------------------

def load_tables():
    wb = load_workbook(EXCEL_FILE)

    def load_wood_table():
        ws = wb["All Woods"]
        mapping = {}
        eyes, traits, paths = set(), set(), set()

        for row in ws.iter_rows(min_row=3, values_only=True):
            eye, trait, path, wood, _note = row
            if not eye or not trait or not path or not wood:
                continue
            eye = eye.strip()
            trait = trait.strip()
            path = path.strip()
            wood = wood.strip()
            mapping[(eye, trait, path)] = wood
            eyes.add(eye)
            traits.add(trait)
            paths.add(path)

        return mapping, sorted(eyes), sorted(traits), sorted(paths)

    def load_core_table():
        ws = wb["Cores"]
        rows = list(ws.iter_rows(values_only=True))
        core_map = {}
        artefacts = []

        header_row_idx = None
        for i, row in enumerate(rows):
            if row[0] and isinstance(row[0], str) and row[0].strip().lower() == "artefact":
                header_row_idx = i
                break

        for row in rows[header_row_idx + 1:]:
            artefact = row[0]
            if not artefact:
                continue
            artefact = artefact.strip()
            if artefact not in artefacts:
                artefacts.append(artefact)

            fear_values = row[1:6]
            for fear_index, core in enumerate(fear_values):
                if core:
                    core_map[(artefact, fear_index)] = core.strip()

        return core_map, artefacts

    def load_lengths_table():
        ws = wb["Lengths"]
        rows = list(ws.iter_rows(values_only=True))

        header_row_idx = None
        for i, row in enumerate(rows):
            if row[0] and isinstance(row[0], str) and row[0].strip().lower() == "artefact":
                header_row_idx = i
                break

        header = rows[header_row_idx]
        categories = [c for c in header[1:] if c]

        length_map = {}
        for row in rows[header_row_idx + 1:]:
            artefact = row[0]
            if not artefact:
                continue
            artefact = artefact.strip()
            for i, cat in enumerate(categories, start=1):
                length = row[i]
                if length:
                    length_map[(artefact, cat.strip().lower())] = str(length).strip()

        return length_map

    def load_flex_table():
        ws = wb["Flexibilities"]
        flex_map = {}

        for row in ws.iter_rows(values_only=True):
            trait = row[0]
            even_flex = row[1]
            odd_flex = row[2]
            if not trait or not even_flex or not odd_flex:
                continue
            if isinstance(trait, str) and trait.strip().lower().startswith("most proud of"):
                continue

            trait = trait.strip()
            flex_map[trait] = {
                "even": str(even_flex).strip(),
                "odd": str(odd_flex).strip(),
            }

        return flex_map

    wood_table, eyes, traits, paths = load_wood_table()
    core_table, artefacts = load_core_table()
    length_table = load_lengths_table()
    flex_table = load_flex_table()

    return {
        "wood_table": wood_table,
        "eyes": eyes,
        "traits": traits,
        "paths": paths,
        "core_table": core_table,
        "artefacts": artefacts,
        "length_table": length_table,
        "flex_table": flex_table
    }

# -----------------------------
# MAIN QUIZ LOGIC (no UI)
# -----------------------------

def compute_wand(answers, tables):
    """
    answers = dict with:
        - eye
        - trait
        - path
        - artefact
        - fear_index
        - height_cat   ('short'/'average'/'tall')
        - parity       ('even'/'odd')

    Returns dict with:
        wood, core, length, flexibility
    """

    wood = tables["wood_table"].get(
        (answers["eye"], answers["trait"], answers["path"])
    )

    core = tables["core_table"].get(
        (answers["artefact"], answers["fear_index"])
    )

    length = tables["length_table"].get(
        (answers["artefact"], answers["height_cat"])
    )

    trait = answers["trait"]
    parity = answers["parity"]

    flex = None
    if trait in tables["flex_table"]:
        flex = tables["flex_table"][trait][parity]

    return {
        "wood": wood,
        "core": core,
        "length": length,
        "flexibility": flex
    }

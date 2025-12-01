#!/usr/bin/env python3

from openpyxl import load_workbook

EXCEL_FILE = "Pottermore Wand Selection Spreadsheet.xlsx"


# -----------------------------
# Utility: asking the user
# -----------------------------

def ask_menu(prompt, options):
    """
    Show a numbered menu from a list of option labels.
    Returns the chosen label (string).
    """
    print(prompt)
    for i, opt in enumerate(options, start=1):
        print(f"    ({i}) {opt}")
    while True:
        choice = input("> ").strip()
        if not choice.isdigit():
            print("Please enter a number.")
            continue
        idx = int(choice)
        if 1 <= idx <= len(options):
            return options[idx - 1]
        print("Please choose a valid option.")


def ask_int(prompt, min_value=None, max_value=None):
    """
    Ask the user for an integer, optionally with bounds.
    """
    while True:
        raw = input(prompt).strip()
        if not raw.isdigit():
            print("Please enter a valid number.")
            continue
        value = int(raw)
        if min_value is not None and value < min_value:
            print(f"Please enter a number >= {min_value}.")
            continue
        if max_value is not None and value > max_value:
            print(f"Please enter a number <= {max_value}.")
            continue
        return value


# -----------------------------
# Loading data from the Excel file
# -----------------------------

def load_wood_table(wb):
    """
    From 'All Woods':
        Eye colour | Trait | Path | Wood | (notes/forbidden)
    Returns mapping (eye, trait, path) -> wood
    and also the sets of available eyes & traits.
    """
    ws = wb["All Woods"]
    mapping = {}
    eyes = set()
    traits = set()
    paths = set()

    # Skip first two rows (description + header)
    for row in ws.iter_rows(min_row=3, values_only=True):
        eye, trait, path, wood, _note = row
        if not eye or not trait or not path or not wood:
            continue
        eye = eye.strip()
        trait = trait.strip()
        path = path.strip()
        wood = wood.strip()
        mapping[(eye, trait, path)] = wood
        eyes.add(eye)
        traits.add(trait)
        paths.add(path)

    return mapping, sorted(eyes), sorted(traits), sorted(paths)


def load_core_table(wb):
    """
    From 'Cores':
      Row ~3: ['Artefact', <fear1>, <fear2>, <fear3>, <fear4>, <fear5>, ...]
      Following rows: artefact + 5 core values
    We'll index fears by 0..4.
    Returns: mapping (artefact, fear_index) -> core
    and a list of artefacts actually present.
    """
    ws = wb["Cores"]
    rows = list(ws.iter_rows(values_only=True))
    core_map = {}
    artefacts = []

    # Find header row that starts with 'Artefact'
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[0] and isinstance(row[0], str) and row[0].strip().lower() == "artefact":
            header_row_idx = i
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find 'Artefact' header row in Cores sheet.")

    # Data rows follow
    for row in rows[header_row_idx + 1:]:
        artefact = row[0]
        if not artefact:
            continue
        artefact = artefact.strip()
        if artefact not in artefacts:
            artefacts.append(artefact)
        # Fear columns: next 5 columns (index 1..5)
        fear_values = row[1:6]
        for fear_index, core in enumerate(fear_values):
            if core:
                core_map[(artefact, fear_index)] = core.strip()

    return core_map, artefacts


def load_lengths_table(wb):
    """
    From 'Lengths':
      Row ~3: ['Artefact', 'short', 'average', 'tall']
      Data rows: artefact + lengths
    Returns: mapping (artefact, height_cat) -> length_str
    """
    ws = wb["Lengths"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row with 'Artefact'
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[0] and isinstance(row[0], str) and row[0].strip().lower() == "artefact":
            header_row_idx = i
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find 'Artefact' header row in Lengths sheet.")

    header = rows[header_row_idx]
    # We'll expect something like: Artefact | short | average | tall
    categories = [c for c in header[1:] if c]

    length_map = {}
    for row in rows[header_row_idx + 1:]:
        artefact = row[0]
        if not artefact:
            continue
        artefact = artefact.strip()
        for i, cat in enumerate(categories, start=1):
            length = row[i]
            if length:
                length_map[(artefact, cat.strip().lower())] = str(length).strip()

    return length_map


def load_flex_table(wb):
    """
    From 'Flexibilities':
      Row ~3: ['most proud of...', 'born on an even date', 'born on an odd date', ...]
      Following rows: Trait | flex_even | flex_odd
    Returns: mapping trait -> {'even': flex_even, 'odd': flex_odd}
    """
    ws = wb["Flexibilities"]
    flex_map = {}

    for row in ws.iter_rows(values_only=True):
        trait = row[0]
        even_flex = row[1]
        odd_flex = row[2]
        if not trait or not even_flex or not odd_flex:
            continue
        if isinstance(trait, str) and trait.strip().lower().startswith("most proud of"):
            # header row
            continue
        trait = trait.strip()
        flex_map[trait] = {
            "even": str(even_flex).strip(),
            "odd": str(odd_flex).strip(),
        }

    return flex_map


# -----------------------------
# Quiz logic
# -----------------------------

def determine_height_category():
    """
    Ask user for approximate height and map to 'short', 'average', 'tall'.
    """
    print("\nFirst, let's work out your wand's length...")
    print("Roughly how tall are you?")
    options = [
        "Short – under about 160cm / 5'3\"",
        "Average – about 160–180cm / 5'3\"–5'11\"",
        "Tall – over about 180cm / 5'11\"",
    ]
    choice = ask_menu("", options)
    if choice.startswith("Short"):
        return "short"
    elif choice.startswith("Average"):
        return "average"
    else:
        return "tall"


def determine_birth_parity():
    """
    Ask for birth day of month, return 'even' or 'odd'.
    """
    print("\nNow, when were you born?")
    day = ask_int("Enter the DAY of the month you were born (1–31): ", 1, 31)
    return "even" if day % 2 == 0 else "odd"


def determine_fear_index():
    """
    Ask for greatest fear; return index 0..4 that matches Cores sheet order:
      0: darkness
      1: fire
      2: heights
      3: small spaces
      4: isolation
    """
    print("\nTell me about your greatest fear...")
    fears = [
        "Darkness",
        "Fire",
        "Heights",
        "Small spaces",
        "Isolation / being alone",
    ]
    choice = ask_menu("Which of these frightens you the most?", fears)
    return fears.index(choice)  # 0..4


def determine_path():
    """
    Ask about the kind of place you're drawn to: Sea, Forest, or Castle.
    """
    print("\nYou find yourself at a crossroads in a dream...")
    paths = [
        "Sea – a wild empty shoreline and crashing waves",
        "Forest – a deep, ancient woodland full of secrets",
        "Castle – towering stone walls and hidden corridors",
    ]
    choice = ask_menu("Where do you feel most drawn to?", paths)
    if choice.startswith("Sea"):
        return "Sea"
    elif choice.startswith("Forest"):
        return "Forest"
    else:
        return "Castle"


def determine_height_label_for_lengths(height_cat):
    """
    Map our internal category to what the Lengths sheet header uses.
    Likely: 'short', 'average', 'tall' (already lowercased).
    """
    return height_cat.lower()


def main():
    print("Welcome to the Wand Selection Quiz!\n")
    print("This quiz will use your answers, together with data from the original")
    print("Pottermore-based spreadsheet, to determine your wand's:")
    print("  • Wood")
    print("  • Core")
    print("  • Length")
    print("  • Flexibility")
    print()

    # Load workbook and tables
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"Error: Could not find {EXCEL_FILE}.")
        print("Make sure the Excel file is in the same folder as this script,")
        print("or change the EXCEL_FILE variable at the top.")
        return

    wood_table, eye_options, trait_options, path_options_from_sheet = load_wood_table(wb)
    core_table, artefact_options_raw = load_core_table(wb)
    length_table = load_lengths_table(wb)
    flex_table = load_flex_table(wb)

    # -----------------------------
    # Quiz questions
    # -----------------------------

    # Eye colour
    print("\nFirst, your eye colour...")
    # Nice ordering: put 'Other' at the end if it's present
    eye_opts = [e for e in eye_options if e.lower() != "other"]
    if "Other" in eye_options:
        eye_opts.append("Other")
    eye = ask_menu("What is your natural eye colour?", eye_opts)

    # Trait (most proud of...)
    print("\nNext, what are you most proud of in your character?")
    # We'll show the traits as-is from the sheet:
    trait = ask_menu("Choose the quality that best fits you:", trait_options)

    # Path (Sea / Forest / Castle)
    path = determine_path()

    # Artefact from the trunk (for core + length)
    print("\nImagine a dusty old trunk. Inside are several curious artefacts...")
    # We'll show them in a fixed nice order if we recognize the standard ones,
    # otherwise we'll fall back to the order from the sheet.
    canonical_order = [
        "dusty bottle",
        "old black glove",
        "golden key",
        "bound-up scroll",
        "glittering jewel",
        "silver dagger",
        "ornate mirror",
    ]

    artefact_opts = [a for a in canonical_order if a in artefact_options_raw]
    # In case there are extras or slightly different strings:
    for a in artefact_options_raw:
        if a not in artefact_opts:
            artefact_opts.append(a)

    artefact = ask_menu(
        "Which artefact are you most drawn to?",
        artefact_opts
    )

    # Greatest fear (for core)
    fear_index = determine_fear_index()

    # Height (for length)
    height_cat = determine_height_category()
    length_label = determine_height_label_for_lengths(height_cat)

    # Birth parity (for flexibility)
    parity = determine_birth_parity()

    # -----------------------------
    # Determine wand characteristics
    # -----------------------------

    # Wood
    wood = wood_table.get((eye, trait, path))
    # Core
    core = core_table.get((artefact, fear_index))
    # Length
    length = length_table.get((artefact, length_label))
    # Flexibility
    flex_data = flex_table.get(trait)
    if flex_data:
        flexibility = flex_data[parity]
    else:
        flexibility = None

    # -----------------------------
    # Output result
    # -----------------------------

    print("\n" + "=" * 60)
    print("                  YOUR WAND HAS BEEN CHOSEN")
    print("=" * 60 + "\n")

    if length and wood and core:
        print(f"Wood:        {wood}")
        print(f"Core:        {core}")
        print(f"Length:      {length}")
    else:
        print("We ran into a small problem matching all of your answers to the data.")
        print("Here's what we could determine:")

        if wood:
            print(f"  • Wood: {wood}")
        else:
            print("  • Wood: (could not determine from the spreadsheet)")

        if core:
            print(f"  • Core: {core}")
        else:
            print("  • Core: (could not determine from the spreadsheet)")

        if length:
            print(f"  • Length: {length}")
        else:
            print("  • Length: (could not determine from the spreadsheet)")

    if flexibility:
        print(f"Flexibility: {flexibility}")
    else:
        print("Flexibility: (could not determine from the spreadsheet)")

    print("\nSummary:")
    desc_parts = []
    if length:
        desc_parts.append(length)
    if wood:
        desc_parts.append(f"{wood} wood")
    if core:
        desc_parts.append(f"with a {core} core")
    if flexibility:
        desc_parts.append(f"— {flexibility}")

    if desc_parts:
        print("  " + " ".join(desc_parts))
    else:
        print("  Your wand is mysterious and defies easy description...")

    print("\nThank you for taking the wand quiz!")


if __name__ == "__main__":
    main()
